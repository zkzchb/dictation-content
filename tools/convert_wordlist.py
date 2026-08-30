#!/usr/bin/env python3
"""词表 xlsx → content-pack JSON。

按设计文档 §3/§4：
  * category 只存 4 种粗类 {生字, 词语, 易错字, 多音字}
  * vocab 的细分 vtype、typo 的配对 pair_id 一并放进 options_json
  * 多音字按 ppid 分组，每组一条 kp，options_json 含各读音
  * 生字保留 word1/word2 两个候选，出题时随机取一个

用法:
  python tools/convert_wordlist.py wordlist_template.xlsx
  python tools/convert_wordlist.py wordlist_template.xlsx \
      --pack-root packs/zh-cn/primary-3a

产出:
  <pack-root>/lessons.json
  <pack-root>/knowledge_points.json
  <pack-root>/studio_manifest.json

该工具只生成结构化草稿；发布前还必须补充稳定 ID、dataset.json、音频和
tts.sha256，并用程序仓库的 content-pack 校验器完整验证。
"""
import os
import sys
import json
import hashlib
import collections
import argparse

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, ".."))

CAT_CHAR = "生字"
CAT_WORD = "词语"
CAT_TYPO = "易错字"
CAT_POLY = "多音字"


def read_sheet(wb, name):
    """读一张表，过滤空行与注释行（首列非纯数字）。"""
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for raw in rows[1:]:
        if raw is None or raw[0] is None:
            continue
        first = str(raw[0]).strip()
        if not first.isdigit():
            continue
        rec = {}
        for i, key in enumerate(header):
            if not key:
                continue
            val = raw[i] if i < len(raw) else None
            rec[key] = str(val).strip() if val is not None else ""
        out.append(rec)
    return out


def word_hash(text):
    return hashlib.md5(text.encode("utf-8")).hexdigest()[:12]


def main():
    parser = argparse.ArgumentParser(description="把词表工作簿转换为 content-pack JSON 草稿")
    parser.add_argument("xlsx", nargs="?", default=os.path.join(ROOT, "wordlist_template.xlsx"))
    parser.add_argument(
        "--pack-root",
        default=os.path.join(ROOT, "packs", "zh-cn", "primary-3a"),
        help="输出内容包目录（默认 packs/zh-cn/primary-3a）",
    )
    args = parser.parse_args()
    xlsx = args.xlsx
    pack_root = os.path.abspath(args.pack_root)
    xlsx = os.path.abspath(xlsx)
    if not os.path.exists(xlsx):
        print(f"[X] 找不到词表: {xlsx}")
        sys.exit(1)

    wb = load_workbook(xlsx, data_only=True)
    units = read_sheet(wb, "unit")
    lessons = read_sheet(wb, "lesson")
    w2w = read_sheet(wb, "word2write")
    vocab = read_sheet(wb, "vocab")
    typo = read_sheet(wb, "typo")
    poly = read_sheet(wb, "polyphonic")

    print(f"读取: {len(units)} 单元 / {len(lessons)} 课 / {len(w2w)} 生字 / "
          f"{len(vocab)} 词语 / {len(typo)} 易错 / {len(poly)} 多音字行")

    # ── 课程目录 ────────────────────────────────────────────────────────
    unit_name = {int(u["uid"]): u.get("utitle", "") for u in units if u.get("uid")}
    lessons_out = []
    for r in lessons:
        lid = int(r["lid"])
        uid = lid // 10
        lessons_out.append({
            "lesson_seq": lid,
            "unit_id": uid,
            "unit_name": unit_name.get(uid, f"单元{uid}"),
            # ltitle 是「第1课」「语文园地」「第一单元」这类序号标题，
            # 前端下拉菜单要靠它显示成「第1课 - 大青树下的小学」。
            "lesson_title": r.get("ltitle", "") or "",
            "lesson_name": r.get("lname", "") or r.get("ltitle", ""),
        })
    lessons_out.sort(key=lambda x: x["lesson_seq"])

    kps = []
    # text -> pinyin，用于录音清单（去重）
    audio_needed = {}

    def need_audio(text, pinyin):
        if text and text not in audio_needed:
            audio_needed[text] = pinyin or ""

    # ── 生字：保留两个组词候选 ───────────────────────────────────────────
    for r in w2w:
        lid = int(r.get("lid") or (int(r["wid"]) // 100))
        char = r.get("basic_word", "")
        if not char:
            continue
        opts = []
        for i in (1, 2):
            w = r.get(f"word{i}", "")
            p = r.get(f"word{i}py", "")
            if w:
                opts.append({"text": w, "pinyin": p})
                need_audio(w, p)
        if not opts:
            # 兜底：没有组词就听写单字本身
            opts = [{"text": char, "pinyin": ""}]
            need_audio(char, "")
        kps.append({
            "lesson_seq": lid, "target": char,
            "category": CAT_CHAR, "options_json": opts,
        })

    # ── 词语：vtype 需向下继承（分组只标首行）───────────────────────────
    cur_vtype = ""
    for r in vocab:
        vt = r.get("vtype", "")
        if vt:
            cur_vtype = vt
        word = r.get("vword", "")
        if not word:
            continue
        lid = int(r.get("lid") or (int(r["vid"]) // 100))
        py = r.get("vpy", "")
        need_audio(word, py)
        kps.append({
            "lesson_seq": lid, "target": word,
            "category": CAT_WORD,
            "options_json": [{"text": word, "pinyin": py, "vtype": cur_vtype or "词语"}],
        })

    # ── 易错字：tw1/tw2 各成一条，用 pair_id 保持配对 ────────────────────
    for r in typo:
        tid = r.get("tid", "")
        lid = int(r.get("lid") or (int(tid) // 100))
        for slot, wcol, pcol in (("tw1", "tw1", "tw1py"), ("tw2", "tw2", "tw2py")):
            word = r.get(wcol, "")
            py = r.get(pcol, "")
            if not word:
                continue
            need_audio(word, py)
            kps.append({
                "lesson_seq": lid, "target": word,
                "category": CAT_TYPO,
                "options_json": [{
                    "text": word, "pinyin": py,
                    "pair_id": tid, "role": slot,
                    "confuse_char": r.get("typo_word", ""),
                }],
            })

    # ── 多音字：按 ppid 分组，一组一条 kp ────────────────────────────────
    groups = collections.OrderedDict()
    for r in poly:
        ppid = r.get("ppid", "")
        if ppid:
            groups.setdefault(ppid, []).append(r)
    for ppid, rows in groups.items():
        char = rows[0].get("pw", "")
        if not char:
            continue
        lid = int(rows[0].get("lid") or (int(ppid) // 100))
        opts = []
        for r in rows:
            w = r.get("word", "")
            if not w:
                continue
            opts.append({
                "text": w,
                "pinyin": r.get("word_py", ""),
                "pron": r.get("pron", ""),
            })
        if not opts:
            continue
        # 多音字只需播报"字"本身的音频，例词仅作参考不进清单
        need_audio(char, "")
        kps.append({
            "lesson_seq": lid, "target": char,
            "category": CAT_POLY, "options_json": opts,
        })

    # ── 录音清单 ────────────────────────────────────────────────────────
    audio_dir = os.path.join(pack_root, "tts", "w")
    recorded = set()
    if os.path.isdir(audio_dir):
        recorded = {f[:-4] for f in os.listdir(audio_dir) if f.endswith(".mp3")}

    manifest = [{"text": t, "pinyin": p, "hash": word_hash(t)}
                for t, p in audio_needed.items()]
    manifest.sort(key=lambda x: x["text"])
    pending = [m for m in manifest if m["hash"] not in recorded]

    # ── 写出 ────────────────────────────────────────────────────────────
    os.makedirs(pack_root, exist_ok=True)
    p_lessons = os.path.join(pack_root, "lessons.json")
    p_kp = os.path.join(pack_root, "knowledge_points.json")
    p_manifest = os.path.join(pack_root, "studio_manifest.json")

    with open(p_lessons, "w", encoding="utf-8") as f:
        json.dump(lessons_out, f, ensure_ascii=False, indent=2)
    with open(p_kp, "w", encoding="utf-8") as f:
        json.dump(kps, f, ensure_ascii=False, indent=2)
    with open(p_manifest, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    by_cat = collections.Counter(k["category"] for k in kps)
    print(f"[OK] {p_lessons}  ({len(lessons_out)} 课)")
    print(f"[OK] {p_kp}  ({len(kps)} 知识点: " +
          ", ".join(f"{c}={n}" for c, n in by_cat.items()) + ")")
    print(f"[OK] {p_manifest}  ({len(manifest)} 唯一词条, {len(pending)} 待录音)")


if __name__ == "__main__":
    main()
