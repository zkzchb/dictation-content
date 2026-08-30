"""批量补全词表拼音（只填空白，不覆盖已有值）。

用法：
  python shared/tools/fill_pinyin.py wordlist_template.xlsx

处理范围：
  word2write  word1py / word2py
  vocab       vpy
  typo        tw1py / tw2py
  polyphonic  pron / word_py

多音字策略：
  * lesson 3000 冷启动的 13 组多音字，读音由 COLD_POLY 人工指定（pypinyin 易错）；
  * 其余表格用 pypinyin 生成，凡含高风险多音字或轻声尾字的词都写入复核清单。

产出：
  * 原地更新 xlsx（先自动备份 .bak_pinyin）
  * _pinyin_review.txt —— 需人工核对的词清单
"""
import os
import sys
import shutil

from openpyxl import load_workbook
from pypinyin import pinyin, Style

# ── lesson 3000 冷启动多音字：(ppid, word) → (pron, word_py) ──────────────
# 这些读音按人教版教学惯例人工确定，轻声按教材写法（不标调）。
COLD_POLY = {
    ("300081", "应该"): ("yīng", "yīng gāi"),
    ("300081", "应用"): ("yìng", "yìng yòng"),
    ("300082", "发现"): ("fā",   "fā xiàn"),
    ("300082", "头发"): ("fà",   "tóu fa"),
    ("300083", "漂流"): ("piāo", "piāo liú"),
    ("300083", "漂白"): ("piǎo", "piǎo bái"),
    ("300083", "漂亮"): ("piào", "piào liang"),
    ("300084", "油炸"): ("zhá",  "yóu zhá"),
    ("300084", "炸弹"): ("zhà",  "zhà dàn"),
    ("300085", "羊圈"): ("juàn", "yáng juàn"),
    ("300085", "圆圈"): ("quān", "yuán quān"),
    ("300086", "方便"): ("biàn", "fāng biàn"),
    ("300086", "便宜"): ("pián", "pián yi"),
    ("300087", "模样"): ("mú",   "mú yàng"),
    ("300087", "模型"): ("mó",   "mó xíng"),
    ("300088", "重复"): ("chóng", "chóng fù"),
    ("300088", "重量"): ("zhòng", "zhòng liàng"),
    ("300089", "尽管"): ("jǐn",  "jǐn guǎn"),
    ("300089", "尽力"): ("jìn",  "jìn lì"),
    ("300090", "背包"): ("bēi",  "bēi bāo"),
    ("300090", "后背"): ("bèi",  "hòu bèi"),
    ("300091", "传说"): ("chuán", "chuán shuō"),
    ("300091", "传记"): ("zhuàn", "zhuàn jì"),
    ("300092", "结实"): ("jiē",  "jiē shi"),
    ("300092", "打结"): ("jié",  "dǎ jié"),
    ("300093", "教书"): ("jiāo", "jiāo shū"),
    ("300093", "教师"): ("jiào", "jiào shī"),
}

# 高风险多音字：pypinyin 在词组中较易选错读音，需人工复核
RISKY = set("发着长重行空少好数处分相转种中便曲系干落教差卷华曾"
            "为得的地了过还只当应更漂模尽背传结圈炸朝挑磨舍薄")

# 常见轻声尾字：pypinyin 会标出本调，教材多写轻声
NEUTRAL_TAIL = set("子们头巴么呢吧了的地得着过们个来去上下里边")


def to_py(text: str) -> str:
    """转成空格分隔的带调拼音。"""
    return " ".join(p[0] for p in pinyin(text, style=Style.TONE))


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return str(v).strip() if v is not None else ""


def header_map(ws):
    return {cell(ws, 1, c): c for c in range(1, ws.max_column + 1) if cell(ws, 1, c)}


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "wordlist_template.xlsx"
    path = os.path.abspath(path)
    if not os.path.exists(path):
        print(f"[ERR] 找不到文件: {path}")
        sys.exit(1)

    backup = path + ".bak_pinyin"
    shutil.copy(path, backup)
    print(f"[OK] 已备份 -> {os.path.basename(backup)}")

    wb = load_workbook(path)
    review = []          # (表, id, 词, 生成拼音, 原因)
    counts = {}

    # ── 1. word2write: word1py / word2py ────────────────────────────────
    ws = wb["word2write"]
    h = header_map(ws)
    n = 0
    for r in range(2, ws.max_row + 1):
        wid = cell(ws, r, h["wid"])
        if not wid:
            continue
        for wcol, pcol in (("word1", "word1py"), ("word2", "word2py")):
            if wcol not in h or pcol not in h:
                continue
            word = cell(ws, r, h[wcol])
            if not word or cell(ws, r, h[pcol]):
                continue
            py = to_py(word)
            ws.cell(row=r, column=h[pcol]).value = py
            n += 1
            why = []
            if RISKY & set(word):
                why.append("多音字")
            if word[-1] in NEUTRAL_TAIL:
                why.append("轻声尾")
            if why:
                review.append(("word2write", wid, word, py, "/".join(why)))
    counts["word2write"] = n

    # ── 2. vocab: vpy ───────────────────────────────────────────────────
    ws = wb["vocab"]
    h = header_map(ws)
    n = 0
    for r in range(2, ws.max_row + 1):
        vid = cell(ws, r, h["vid"])
        word = cell(ws, r, h["vword"])
        if not vid or not word or cell(ws, r, h["vpy"]):
            continue
        py = to_py(word)
        ws.cell(row=r, column=h["vpy"]).value = py
        n += 1
        why = []
        if RISKY & set(word):
            why.append("多音字")
        if word[-1] in NEUTRAL_TAIL:
            why.append("轻声尾")
        if why:
            review.append(("vocab", vid, word, py, "/".join(why)))
    counts["vocab"] = n

    # ── 3. typo: tw1py / tw2py ──────────────────────────────────────────
    ws = wb["typo"]
    h = header_map(ws)
    n = 0
    for r in range(2, ws.max_row + 1):
        tid = cell(ws, r, h["tid"])
        if not tid:
            continue
        for wcol, pcol in (("tw1", "tw1py"), ("tw2", "tw2py")):
            word = cell(ws, r, h[wcol])
            if not word or cell(ws, r, h[pcol]):
                continue
            py = to_py(word)
            ws.cell(row=r, column=h[pcol]).value = py
            n += 1
            why = []
            if RISKY & set(word):
                why.append("多音字")
            if word[-1] in NEUTRAL_TAIL:
                why.append("轻声尾")
            if why:
                review.append(("typo", tid, word, py, "/".join(why)))
    counts["typo"] = n

    # ── 4. polyphonic: pron / word_py（含 较→教 修正）────────────────────
    ws = wb["polyphonic"]
    h = header_map(ws)
    n = 0
    fixed_pw = 0
    for r in range(2, ws.max_row + 1):
        ppid = cell(ws, r, h["ppid"])
        word = cell(ws, r, h["word"])
        if not ppid or not word:
            continue

        # 笔误修正：ppid 300093 的 pw 应为「教」
        if ppid == "300093" and cell(ws, r, h["pw"]) == "较":
            ws.cell(row=r, column=h["pw"]).value = "教"
            fixed_pw += 1

        key = (ppid, word)
        if key in COLD_POLY:
            pron, wpy = COLD_POLY[key]
            if not cell(ws, r, h["pron"]):
                ws.cell(row=r, column=h["pron"]).value = pron
            if not cell(ws, r, h["word_py"]):
                ws.cell(row=r, column=h["word_py"]).value = wpy
                n += 1
            continue

        # 非冷启动组：pypinyin 兜底，并从词中定位本字读音
        if not cell(ws, r, h["word_py"]):
            wpy = to_py(word)
            ws.cell(row=r, column=h["word_py"]).value = wpy
            n += 1
            review.append(("polyphonic", ppid, word, wpy, "多音字-需确认读音"))
            if not cell(ws, r, h["pron"]):
                pw = cell(ws, r, h["pw"])
                syl = wpy.split()
                idx = word.find(pw) if pw else -1
                if 0 <= idx < len(syl):
                    ws.cell(row=r, column=h["pron"]).value = syl[idx]
    counts["polyphonic"] = n

    wb.save(path)

    # ── 复核清单 ────────────────────────────────────────────────────────
    out = os.path.join(os.path.dirname(path), "_pinyin_review.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write("需人工核对的拼音（含多音字或轻声尾字）\n")
        f.write("=" * 60 + "\n")
        for tbl, wid, word, py, why in review:
            f.write(f"{tbl:12} {wid:8} {word:10} {py:28} [{why}]\n")

    print(f"[OK] 拼音补全完成，共 {sum(counts.values())} 处")
    for k, v in counts.items():
        print(f"     {k}: {v}")
    if fixed_pw:
        print(f"[OK] polyphonic 300093 的 pw 已由「较」修正为「教」（{fixed_pw} 行）")
    print(f"[OK] 复核清单 {len(review)} 条 -> {os.path.basename(out)}")


if __name__ == "__main__":
    main()
